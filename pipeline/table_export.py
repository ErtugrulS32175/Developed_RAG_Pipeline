import csv
import json
import re
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline.text_normalize import has_residual_marks

# Highlight for cells a reviewer should check (consensus disagreements): light amber.
_REVIEW_FILL = PatternFill("solid", fgColor="FFE699")


def _int_attr(attrs, name, default=1):
    """Parse an HTML integer attribute (colspan/rowspan) tolerantly; VLM OCR
    sometimes emits stray/non-numeric values, so fall back to `default` (1) and
    clamp to [1, 1000] -- a garbled span like rowspan="8000" would otherwise
    balloon the grid's occupied-cell set. No real table span approaches 1000."""
    for k, v in attrs:
        if k == name:
            try:
                return min(1000, max(1, int(str(v).strip())))
            except (TypeError, ValueError):
                return default
    return default


class _HTMLTableExtractor(HTMLParser):
    """Pull <table>s out of a VLM's markdown/HTML output into rows of cell text,
    remembering which rows sit inside <thead>. Tolerant of the messy HTML
    doc-parsing VLMs emit: ignores unknown tags, treats <th>/<td> as cells, <tr>
    as rows. Cell spans (colspan/rowspan) are NOT expanded -- but a full-width
    spanning TITLE row (e.g. <th colspan=14>REPORT TITLE</th>) lands in
    <thead> with few cells, so parse_html_tables can skip it via the head flags."""

    def __init__(self):
        super().__init__()
        self.tables = []                       # list of (rows, head_flags, spans)
        self._rows = self._flags = self._spans = self._cell = None
        self._span = (1, 1)
        self._in_thead = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._rows, self._flags, self._spans = [], [], []
            self._in_thead = False
        elif tag == "thead" and self._rows is not None:
            self._in_thead = True
        elif tag == "tbody" and self._rows is not None:
            self._in_thead = False
        elif tag == "tr" and self._rows is not None:
            self._rows.append([])
            self._spans.append([])
            self._flags.append(self._in_thead)
        elif tag in ("td", "th") and self._rows:
            self._cell = []
            self._span = (_int_attr(attrs, "colspan"), _int_attr(attrs, "rowspan"))

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._cell is not None:
            self._rows[-1].append(" ".join("".join(self._cell).split()))
            self._spans[-1].append(self._span)
            self._cell = None
        elif tag == "table":
            self._commit()

    def _commit(self):
        """Emit the current table's non-empty rows. Called on </table> AND at
        end-of-feed, so a VLM output truncated at its token cap (rows present but
        no closing </table>) still yields the rows it managed to produce."""
        if self._cell is not None:            # truncated mid-cell: flush partial
            self._rows[-1].append(" ".join("".join(self._cell).split()))
            self._spans[-1].append(self._span)
            self._cell = None
        if self._rows:
            keep = [(r, f, s) for r, f, s in zip(self._rows, self._flags, self._spans) if r]
            if keep:
                self.tables.append(
                    ([r for r, _, _ in keep], [f for _, f, _ in keep], [s for _, _, s in keep])
                )
        self._rows = self._flags = self._spans = None


def parse_html_tables(text):
    """VLM markdown/HTML -> [{headers, rows, ...}]. Routes each table by shape:

      * flat (single-row header) -> _parse_flat: header is the widest <thead>
        row, or without <thead> the first full-grid row (leading spanning TITLE
        rows are skipped). This covers the common case and title-only colspans.
      * grouped (a rowspan>1 is present) -> _parse_grouped: a real two-level
        header. Spans are expanded, the header band folded into flat "Group -
        Sub" column names, and the result also carries `header_rows` /
        `header_merges` so the Excel exporter can rebuild the merged header.

    Every result has `headers` and `rows`; grouped ones add the two extra keys.
    Returns [] if no <table> is present."""
    p = _HTMLTableExtractor()
    try:
        p.feed(text or "")
    except Exception:
        pass
    p._commit()          # flush a table left open by truncated (token-capped) output
    out = []
    for rows, flags, spans in p.tables:
        if not rows:
            continue
        # A rowspan>1 anywhere is the signal for a genuine multi-row (grouped)
        # header -- a group label sits above sub-columns and spans down into the
        # sub-label row. Colspan alone (a full-width TITLE row) does NOT count;
        # that stays on the flat path so title-only tables keep working.
        if any(rs > 1 for row in spans for _, rs in row):
            out.append(_parse_grouped(rows, spans))
        else:
            out.append(_parse_flat(rows, flags))
    return out


def _parse_flat(rows, flags):
    """Single-row-header parse (no grouped header). The column header is the
    WIDEST <thead> row when <thead> is present; without <thead> it's the first
    row that spans the full grid (leading spanning TITLE rows, which collapse to
    fewer cells, are skipped)."""
    head_idx = [i for i, f in enumerate(flags) if f]
    if head_idx:
        hi = max(head_idx, key=lambda i: (len(rows[i]), i))   # widest thead row
        headers = rows[hi]
        body = [r for i, r in enumerate(rows) if not flags[i]]
    else:
        # No <thead>: the header is the first row that spans the full grid. A
        # spanning TITLE/caption row above it collapses to fewer cells (often
        # one), so skip leading rows narrower than the table's dominant width.
        width = Counter(len(r) for r in rows).most_common(1)[0][0]
        start = 0
        while start < len(rows) - 1 and len(rows[start]) < width:
            start += 1
        headers = rows[start]
        body = rows[start + 1:]
    return {"headers": headers, "rows": body}


def _build_grid(rows, spans):
    """Expand colspan/rowspan into a rectangular grid using the HTML table
    layout algorithm. Returns (grid, labels, merges, width): `grid` places each
    cell's text at its top-left and blanks the covered cells (what Excel shows
    under a merge); `labels` fills EVERY covered cell with the text (used to fold
    a grouped header into flat column names); `merges` lists (r, c, rowspan,
    colspan) for every spanning cell."""
    grid, labels, merges, occupied = {}, {}, [], set()
    width = 0
    for r, (cells, cellspans) in enumerate(zip(rows, spans)):
        c = 0
        for text, (cs, rs) in zip(cells, cellspans):
            while (r, c) in occupied:
                c += 1
            for dr in range(rs):
                for dc in range(cs):
                    occupied.add((r + dr, c + dc))
                    grid[(r + dr, c + dc)] = text if (dr == 0 and dc == 0) else ""
                    labels[(r + dr, c + dc)] = text
            if cs > 1 or rs > 1:
                merges.append((r, c, rs, cs))
            c += cs
            width = max(width, c)
    return grid, labels, merges, width


def flatten_header(header_rows, header_merges):
    """Fold a two-level header (a display grid where merged cells hold their text
    at the top-left and blanks elsewhere, plus the merge spans) into one flat
    "Group - Sub" name per column. Used both by the grouped parser and by the
    template stamper so both produce identical flat column names."""
    if not header_rows:
        return []
    n = len(header_rows)
    width = max(len(r) for r in header_rows)
    # rebuild the full label coverage: every covered cell carries its span's text
    label = {}
    for r in range(n):
        for c in range(len(header_rows[r])):
            if str(header_rows[r][c]).strip():
                label[(r, c)] = header_rows[r][c]
    for (r, c, rs, cs) in header_merges:
        text = header_rows[r][c] if c < len(header_rows[r]) else ""
        for dr in range(rs):
            for dc in range(cs):
                label[(r + dr, c + dc)] = text
    out = []
    for c in range(width):
        seen = []
        for r in range(n):
            lab = label.get((r, c), "")
            if str(lab).strip() and (not seen or seen[-1] != lab):
                seen.append(str(lab))
        out.append(" - ".join(seen))
    return out


def _parse_grouped(rows, spans):
    """Parse a table with a genuine multi-row grouped header (rowspan present).
    Expands spans to a grid, skips a leading full-width TITLE row, detects how
    many rows the header band covers (from the rowspan reach of its first row),
    folds the band into flat "Group - Sub" column names, and returns the data
    rows below it. Extra keys `header_rows` (the raw header grid) and
    `header_merges` (spans relative to the header block) let the Excel exporter
    reproduce the two-level merged header faithfully."""
    grid, _labels, merges, width = _build_grid(rows, spans)
    nrows = len(rows)

    def row_cells(g, r):
        return [g.get((r, c), "") for c in range(width)]

    # Skip a leading TITLE row: a single cell spanning the full width.
    h0 = 0
    while h0 < nrows - 1:
        row_merges = [m for m in merges if m[0] == h0]
        if len(row_merges) == 1 and row_merges[0][1] == 0 and row_merges[0][3] == width:
            h0 += 1
        else:
            break

    # Header band height = how far the first header row's cells span downward.
    band = max((rs for (r, _, rs, _) in merges if r == h0), default=1)
    band = min(band, nrows - h0)                     # never swallow all rows
    header_end = h0 + band

    header_rows = [row_cells(grid, r) for r in range(h0, header_end)]
    header_merges = [(r - h0, c, rs, cs) for (r, c, rs, cs) in merges if h0 <= r < header_end]
    body = [row_cells(grid, r) for r in range(header_end, nrows)]
    return {
        "headers": flatten_header(header_rows, header_merges),
        "rows": body,
        "header_rows": header_rows,
        "header_merges": header_merges,
    }


def _extract_balanced(text, start):
    """Return the balanced bracket substring starting at text[start] (a '[' or
    '{'), string-aware so brackets inside quoted values don't count."""
    open_ch = text[start]
    close_ch = "]" if open_ch == "[" else "}"
    depth = 0
    in_str = esc = False
    for i in range(start, len(text)):
        c = text[i]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
        elif c == '"':
            in_str = True
        elif c == open_ch:
            depth += 1
        elif c == close_ch:
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def parse_table_json(raw):
    """Parse a VLM's table output into {headers, rows}, tolerant of the mess
    VLMs produce: ``` fences and botched trailing brackets (e.g. ']}]' instead
    of ']]}') that make a strict json.loads return nothing. Strategy: strip
    fences, try strict parse, else pull the headers array and each row array
    out individually, skipping stray braces. Lives client-side on purpose so
    parser fixes don't need a redeploy of the GPU table service."""
    if not raw:
        return {"headers": [], "rows": []}
    text = re.sub(r"```$", "", re.sub(r"^```(?:json)?", "", raw.strip())).strip()

    try:
        data = json.loads(text)
        return {"headers": data.get("headers", []), "rows": data.get("rows", [])}
    except (json.JSONDecodeError, AttributeError):
        pass

    headers = []
    hm = re.search(r'"headers"\s*:\s*\[', text)
    if hm:
        h = _extract_balanced(text, hm.end() - 1)
        if h:
            try:
                headers = json.loads(h)
            except json.JSONDecodeError:
                headers = []

    rows = []
    rm = re.search(r'"rows"\s*:\s*\[', text)
    if rm:
        i, depth = rm.end(), 1
        while i < len(text) and depth > 0:
            c = text[i]
            if c == "[":
                row = _extract_balanced(text, i)
                if row is None:
                    break
                try:
                    rows.append(json.loads(row))
                except json.JSONDecodeError:
                    pass
                i += len(row)
                continue
            if c == "]":
                depth -= 1
            i += 1
    return {"headers": headers, "rows": rows}


# Fold Turkish letters to an ASCII base so the OCR cross-check tolerates the
# expected ı/i, ş/s, ğ/g ... disagreement between two OCR engines: text cells
# come from EasyOCR-tr ("Yıldız") but the page cross-check text comes from the
# latin PaddleOCR recognizer ("Yildiz"). Without this, correct Turkish names get
# falsely flagged as hallucination candidates.
_TR_FOLD = str.maketrans({
    "ı": "i", "İ": "i", "I": "i", "ş": "s", "Ş": "s", "ğ": "g", "Ğ": "g",
    "ç": "c", "Ç": "c", "ö": "o", "Ö": "o", "ü": "u", "Ü": "u",
})


def _squash(s) -> str:
    """Lowercase, fold Turkish diacritics, and strip whitespace/punctuation for
    tolerant matching, so neither formatting differences (1.000,50 vs 1000.50,
    spacing) nor cross-engine Turkish-char disagreement trigger false
    hallucination flags in the OCR cross-check."""
    return re.sub(r"[\s\W_]+", "", str(s).translate(_TR_FOLD).lower())


def validate_table(headers, rows, ocr_text=None):
    """Return (confidence, issues). Detects problems, never corrects them --
    low-confidence tables are flagged for a human, not dropped. When ocr_text
    (the same page's OCR output) is given, cross-check every cell against it:
    values that never appear in the OCR text are hallucination candidates
    (invented cells / dropped letters the normalize layer can't catch)."""
    if not headers or not rows:
        return 0.0, ["bos tablo (header veya satir yok)"]

    issues = []
    width = len(headers)

    bad_width = sum(1 for row in rows if len(row) != width)
    if bad_width:
        issues.append(f"{bad_width} satir header genisligiyle uyusmuyor (sutun sayisi tutmuyor)")

    empty_rows = sum(1 for row in rows if all(str(c).strip() == "" for c in row))
    if empty_rows:
        issues.append(f"{empty_rows} tamamen bos satir")

    unmatched = []
    if ocr_text:
        ocr_norm = _squash(ocr_text)
        for row in rows:
            for cell in row:
                token = str(cell).strip()
                if token and _squash(token) not in ocr_norm:
                    unmatched.append(token)
        if unmatched:
            sample = ", ".join(unmatched[:5])
            issues.append(f"{len(unmatched)} hucre OCR metninde yok (uydurma adayi): {sample}")

    flat = " ".join(str(c) for row in rows for c in row)
    if has_residual_marks(flat) or has_residual_marks(" ".join(str(h) for h in headers)):
        issues.append("cozulmemis Turkce karakter isareti kaldi (normalize eksik)")

    confidence = sum(1 for row in rows if len(row) == width) / len(rows)
    n_cells = sum(len(row) for row in rows) or 1
    if unmatched:
        confidence *= max(0.0, 1 - len(unmatched) / n_cells)
    return round(confidence, 2), issues


def estimate_table_confidence(headers, rows) -> float:
    """Cheap deterministic proxy for extraction quality -- not a model
    confidence. Gemma's table output is generative JSON, not a detection
    model, so there's no calibrated per-cell score to report; this just
    checks the shape came back consistent (every row matches the header
    width), which is exactly the kind of structural regression PaddleOCR's
    table model was dropped for (see the table-module-status notes)."""
    if not headers or not rows:
        return 0.0
    width = len(headers)
    consistent = sum(1 for row in rows if len(row) == width)
    return round(consistent / len(rows), 2)


def table_to_markdown(headers, rows, *, filename=None, page=None, table_id=None, confidence=None) -> str:
    if not headers:
        return ""
    lines = []
    if filename is not None or table_id is not None:
        lines.append(f"Belge: {filename}")
        lines.append(f"Sayfa: {page}")
        lines.append(f"Tablo: {table_id}")
        if confidence is not None:
            lines.append(f"Güven: {confidence:.2f}")
        lines.append("")
    lines.append("| " + " | ".join(str(h) for h in headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        lines.append("| " + " | ".join(str(c) for c in row) + " |")
    return "\n".join(lines)


def save_table_json(table_id, page, headers, rows, confidence, path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    data = {
        "table_id": table_id,
        "page": page,
        "headers": headers,
        "rows": rows,
        "confidence": confidence,
    }
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def save_table_xlsx(headers, rows, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_header(ws, result, review_headers):
    """Write the header into the "Tablo" sheet and return how many Excel rows it
    occupies (0 if there is no header). A grouped result (from _parse_grouped)
    carries `header_rows` + `header_merges`, which are rendered as a faithful
    two-level MERGED header; otherwise a single bold header row is written.
    `review_headers` is the set of column indices to highlight amber.
    `review_all_headers` on the result highlights the WHOLE header -- used when a
    grouped header could not be mapped to a known form ("undefined form")."""
    header_rows = result.get("header_rows")
    headers = result.get("headers", [])
    review_all = bool(result.get("review_all_headers"))

    if header_rows:
        merges = result.get("header_merges", [])
        n = len(header_rows)
        width = max((len(hr) for hr in header_rows), default=0)
        for hr in header_rows:
            ws.append(list(hr))
        for (r, c, rs, cs) in merges:
            # clamp inside the header block so a stray rowspan can't spill into
            # (and later crash `ws.append` on) the data rows below
            end_row = min(r + rs, n)
            end_col = min(c + cs, width)
            if end_row > r + 1 or end_col > c + 1:
                ws.merge_cells(start_row=r + 1, start_column=c + 1,
                               end_row=end_row, end_column=end_col)
        for r in range(1, n + 1):
            for cell in ws[r]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if review_all:
                    cell.fill = _REVIEW_FILL
        # header disagreements are per flat column -> mark the bottom (leaf) row
        for j in review_headers:
            ws.cell(row=n, column=j + 1).fill = _REVIEW_FILL
        return n

    if headers:
        ws.append(list(headers))
        for j, cell in enumerate(ws[1]):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if review_all or j in review_headers:
                cell.fill = _REVIEW_FILL
        return 1

    return 0


def _write_table_sheet(ws, table, review_cells=frozenset(), review_headers=frozenset()):
    """Render one table (flat or two-level grouped header + data) into a sheet:
    bold/frozen/merged header, auto-ish column widths, and any cell in
    `review_cells` / column in `review_headers` highlighted amber."""
    headers = table.get("headers", [])
    rows = table.get("rows", [])

    n_header = _write_header(ws, table, review_headers)
    if n_header:
        ws.freeze_panes = f"A{n_header + 1}"

    for i, row in enumerate(rows):
        ws.append(list(row))
        excel_row = i + 1 + n_header
        for j in range(len(row)):
            if (i, j) in review_cells:
                ws.cell(row=excel_row, column=j + 1).fill = _REVIEW_FILL

    # rough column widths from the longest value seen per column
    for j in range(len(headers) or (len(rows[0]) if rows else 0)):
        longest = len(str(headers[j])) if j < len(headers) else 0
        for row in rows:
            if j < len(row):
                longest = max(longest, len(str(row[j])))
        ws.column_dimensions[get_column_letter(j + 1)].width = min(max(longest + 2, 8), 40)


def _candidate_sheet_title(cand, k):
    """Excel-safe sheet title for one show-both candidate (<=31 chars, no []:*?/\\)."""
    label = cand.get("backend") or f"model{k + 1}"
    title = f"Model {chr(65 + k)} - {label}"
    for ch in r'[]:*?/\ ':
        title = title.replace(ch, "_")
    return title[:31]


def _candidate_width(cand):
    """Column count of a candidate reading (grouped header width if present)."""
    hr = cand.get("header_rows")
    if hr:
        return max((len(r) for r in hr), default=0)
    rows = cand.get("rows", [])
    return len(cand.get("headers", [])) or (len(rows[0]) if rows else 0)


def _write_comparison_sheet(ws, candidates):
    """Decision dashboard for a show-both result: one column per model with its
    shape, header type, and quality signal. Deliberately NOT a cross-model header
    alignment -- the models disagree on column count, so pairing their headers
    cell-by-cell would misalign; each model's full header lives on its own sheet.
    Here we only report per-model summary values (nothing to shift)."""
    labels = [c.get("backend") or f"model{k + 1}" for k, c in enumerate(candidates)]
    rows = [
        ["Karsilastirma", *labels],
        ["Sutun sayisi", *[_candidate_width(c) for c in candidates]],
        ["Satir sayisi", *[len(c.get("rows", [])) for c in candidates]],
        ["Baslik tipi", *["gruplu" if c.get("header_rows") else "duz" for c in candidates]],
        ["Supheli sayi (OCR'da yok)", *[c.get("suspect_count", 0) for c in candidates]],
        ["Sayi guveni", *[c.get("number_fidelity", "") for c in candidates]],
    ]
    for r in rows:
        ws.append(r)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 26
    for k in range(len(candidates)):
        ws.column_dimensions[get_column_letter(k + 2)].width = 22


def export_result_xlsx(result, path):
    """Write a pipeline result dict (from table_pipeline.run or run_consensus) to
    an .xlsx the reviewer can act on:

      * "Tablo" sheet -- bold + frozen header (a two-level MERGED header when the
        result carries header_rows/header_merges from a grouped table), auto-ish
        column widths, and any cell the models disagreed on highlighted amber.
      * BUT when the result carries `candidates` (models disagreed on STRUCTURE and
        no template could arbitrate), a "Karsilastirma" dashboard (per-model shape
        + quality) leads, then each candidate reading is written to its own
        "Model A/B - <backend>" sheet (with that model's suspect cells highlighted)
        so a human can compare and pick.
      * "Rapor" sheet -- backend(s), confidence breakdown, issues, and each
        disagreement with BOTH candidate values so a human can pick.

    Works for a single-backend result (no disagreements -> nothing highlighted)
    and a consensus result alike. Values are written verbatim (no number coercion)
    to preserve OCR fidelity."""
    disagreements = result.get("disagreements", [])
    review_cells = {d["pos"] for d in disagreements if d.get("kind") == "cell"}
    review_headers = {d["pos"] for d in disagreements if d.get("kind") == "header"}

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    candidates = result.get("candidates")
    if candidates:
        # decision dashboard first, then each model's faithful reading on its own
        # sheet with that model's suspect numeric cells highlighted
        _write_comparison_sheet(wb.active, candidates)
        wb.active.title = "Karsilastirma"
        for k, cand in enumerate(candidates):
            ws = wb.create_sheet(_candidate_sheet_title(cand, k))
            _write_table_sheet(ws, cand, review_cells=set(cand.get("review_cells", ())))
    else:
        ws = wb.active
        ws.title = "Tablo"
        _write_table_sheet(ws, result, review_cells, review_headers)

    _write_report_sheet(wb.create_sheet("Rapor"), result)
    wb.save(path)


def _write_report_sheet(ws, result):
    backends = result.get("backends") or ([result["backend"]] if result.get("backend") else [])
    ws.append(["Backend", " + ".join(backends)])
    ws.append(["Guven (confidence)", result.get("confidence")])
    ws.append(["  yapisal", result.get("structural_confidence")])
    ws.append(["  sayisal (number fidelity)", result.get("number_fidelity")])
    if "agreement" in result:
        ws.append(["  model uyumu (agreement)", result.get("agreement")])
    ws.append(["Gozden gecirme gerekli", "EVET" if result.get("needs_review") else "hayir"])
    ws.append([])

    issues = result.get("issues") or []
    ws.append(["Sorunlar", f"{len(issues)} adet"])
    for msg in issues:
        ws.append(["", msg])

    disagreements = result.get("disagreements") or []
    if disagreements:
        ws.append([])
        ws.append(["Ayrisan hucreler", f"{len(disagreements)} adet"])
        ws.append(["konum", "aday 1", "aday 2"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for d in disagreements:
            keys = [k for k in d if k not in ("kind", "pos")]
            pos = d.get("pos")
            loc = (f"satir {pos[0]+1}, sutun {pos[1]+1}" if isinstance(pos, tuple)
                   else f"baslik {pos+1}" if isinstance(pos, int)
                   else d.get("kind", ""))
            ws.append([loc] + [str(d.get(k)) for k in keys])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40


def save_table_csv(headers, rows, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if headers:
            writer.writerow(headers)
        writer.writerows(rows)
