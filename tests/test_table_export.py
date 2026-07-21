from openpyxl import load_workbook

from pipeline.table_export import (
    estimate_table_confidence,
    export_result_xlsx,
    parse_html_tables,
    table_to_markdown,
)


def test_confidence_all_rows_match_header_width():
    headers = ["Product", "Quantity"]
    rows = [["Pen", 5], ["Pencil", 3]]
    assert estimate_table_confidence(headers, rows) == 1.0


def test_confidence_penalizes_mismatched_rows():
    headers = ["Product", "Quantity", "Total"]
    rows = [["Pen", 5, 100], ["Pencil", 3]]  # second row missing a column
    assert estimate_table_confidence(headers, rows) == 0.5


def test_confidence_empty_table_is_zero():
    assert estimate_table_confidence([], []) == 0.0
    assert estimate_table_confidence(["Product"], []) == 0.0


def test_markdown_includes_citation_header_when_provided():
    md = table_to_markdown(
        ["Product", "Quantity"], [["Pen", 5]],
        filename="invoice_001.pdf", page=1, table_id="table_001", confidence=0.94,
    )
    assert "Belge: invoice_001.pdf" in md
    assert "Tablo: table_001" in md
    assert "Güven: 0.94" in md
    assert "| Product | Quantity |" in md


def test_markdown_without_citation_args_stays_plain():
    md = table_to_markdown(["Product"], [["Pen"]])
    assert "Belge:" not in md
    assert md.startswith("| Product |")


def test_parse_skips_spanning_title_row_without_thead():
    html = ("<table>"
            "<tr><td>REPORT TITLE</td></tr>"
            "<tr><td>Product</td><td>Qty</td><td>Total</td></tr>"
            "<tr><td>Pen</td><td>5</td><td>100</td></tr>"
            "</table>")
    t = parse_html_tables(html)[0]
    assert t["headers"] == ["Product", "Qty", "Total"]
    assert t["rows"] == [["Pen", "5", "100"]]


def test_parse_first_row_is_header_when_no_title():
    html = ("<table>"
            "<tr><td>Product</td><td>Qty</td></tr>"
            "<tr><td>Pen</td><td>5</td></tr>"
            "</table>")
    t = parse_html_tables(html)[0]
    assert t["headers"] == ["Product", "Qty"]
    assert t["rows"] == [["Pen", "5"]]


def test_parse_widest_thead_row_is_header():
    html = ("<table><thead>"
            "<tr><th>REPORT TITLE</th></tr>"
            "<tr><th>Product</th><th>Qty</th></tr>"
            "</thead><tbody>"
            "<tr><td>Pen</td><td>5</td></tr>"
            "</tbody></table>")
    t = parse_html_tables(html)[0]
    assert t["headers"] == ["Product", "Qty"]
    assert t["rows"] == [["Pen", "5"]]


def test_parse_recovers_table_truncated_before_close_tag():
    html = "<table><tr><td>Product</td><td>Qty</td></tr><tr><td>Pen</td><td>5</td>"
    t = parse_html_tables(html)[0]
    assert t["headers"] == ["Product", "Qty"]
    assert t["rows"] == [["Pen", "5"]]


def test_parse_colspan_title_without_rowspan_stays_flat():
    # A spanning TITLE row (colspan only, no rowspan) must NOT trigger the
    # grouped-header path: title skipped, the wide row is the flat header.
    html = ("<table>"
            "<tr><td colspan='3'>REPORT TITLE</td></tr>"
            "<tr><td>Product</td><td>Qty</td><td>Total</td></tr>"
            "<tr><td>Pen</td><td>5</td><td>100</td></tr>"
            "</table>")
    t = parse_html_tables(html)[0]
    assert t["headers"] == ["Product", "Qty", "Total"]
    assert t["rows"] == [["Pen", "5", "100"]]
    assert "header_merges" not in t


def test_parse_two_level_grouped_header():
    # rowspan present -> grouped-header path. Group row spans, sub-row labels it.
    #   ColA(cs2,rs2) | ColB(rs2) | GroupC(cs2)
    #                              | Sub1 | Sub2
    html = ("<table>"
            "<tr><td colspan='2' rowspan='2'>ColA</td>"
            "<td rowspan='2'>ColB</td>"
            "<td colspan='2'>GroupC</td></tr>"
            "<tr><td>Sub1</td><td>Sub2</td></tr>"
            "<tr><td>a1</td><td></td><td>b1</td><td>c1</td><td>c2</td></tr>"
            "</table>")
    t = parse_html_tables(html)[0]
    # flat header: group label folded into each sub-column
    assert t["headers"] == ["ColA", "ColA", "ColB", "GroupC - Sub1", "GroupC - Sub2"]
    # the data row is no longer mistaken for the header
    assert t["rows"] == [["a1", "", "b1", "c1", "c2"]]
    # merges are reported (row, col, rowspan, colspan) relative to the header block
    merges = {(m[0], m[1]): (m[2], m[3]) for m in t["header_merges"]}
    assert merges[(0, 0)] == (2, 2)      # ColA: 2 rows x 2 cols
    assert merges[(0, 2)] == (2, 1)      # ColB: 2 rows x 1 col
    assert merges[(0, 3)] == (1, 2)      # GroupC: 1 row x 2 cols


def _consensus_result():
    return {
        "mode": "consensus",
        "backends": ["vl", "hy"],
        "headers": ["Product", "Qty"],
        "rows": [["Pen", "5"], ["Book", "3"]],
        "confidence": 0.83,
        "structural_confidence": 1.0,
        "number_fidelity": 1.0,
        "agreement": 0.83,
        "needs_review": True,
        "issues": ["1 hucrede modeller ayristi"],
        "disagreements": [{"kind": "cell", "pos": (1, 1), "vl": "3", "hy": "8"}],
    }


def test_export_writes_data_and_report_sheets(tmp_path):
    out = tmp_path / "t.xlsx"
    export_result_xlsx(_consensus_result(), str(out))
    wb = load_workbook(out)
    assert wb.sheetnames == ["Tablo", "Rapor"]
    ws = wb["Tablo"]
    assert [c.value for c in ws[1]] == ["Product", "Qty"]
    assert ws["A1"].font.bold is True
    assert ws["B3"].value == "3"                       # row (1,1) -> excel B3


def test_export_highlights_disagreement_cell(tmp_path):
    out = tmp_path / "t.xlsx"
    export_result_xlsx(_consensus_result(), str(out))
    ws = load_workbook(out)["Tablo"]
    assert ws["B3"].fill.patternType == "solid"             # disagreed cell filled
    assert ws["B3"].fill.fgColor.rgb.endswith("FFE699")
    assert ws["A2"].fill.patternType is None                # agreeing cell not filled


def test_export_report_lists_both_candidates(tmp_path):
    out = tmp_path / "t.xlsx"
    export_result_xlsx(_consensus_result(), str(out))
    rows = list(load_workbook(out)["Rapor"].iter_rows(values_only=True))
    flat = [str(c) for r in rows for c in r if c is not None]
    assert "vl + hy" in flat
    assert "3" in flat and "8" in flat                 # both candidates present


def _grouped_result():
    return {
        "backend": "b1",
        "headers": ["ColA", "ColA", "ColB", "GroupC - Sub1", "GroupC - Sub2"],
        "header_rows": [
            ["ColA", "", "ColB", "GroupC", ""],
            ["", "", "", "Sub1", "Sub2"],
        ],
        "header_merges": [(0, 0, 2, 2), (0, 2, 2, 1), (0, 3, 1, 2)],
        "rows": [["a1", "", "b1", "c1", "c2"]],
        "confidence": 1.0,
        "needs_review": False,
        "issues": [],
    }


def test_export_grouped_header_writes_merged_two_level_header(tmp_path):
    out = tmp_path / "g.xlsx"
    export_result_xlsx(_grouped_result(), str(out))
    ws = load_workbook(out)["Tablo"]
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "A1:B2" in merged          # ColA: 2 rows x 2 cols
    assert "C1:C2" in merged          # ColB: 2 rows x 1 col
    assert "D1:E1" in merged          # GroupC: 1 row x 2 cols
    assert ws["A1"].value == "ColA"
    assert ws["C1"].value == "ColB"
    assert ws["D1"].value == "GroupC"
    assert ws["D2"].value == "Sub1" and ws["E2"].value == "Sub2"
    assert ws["A1"].font.bold is True


def test_export_grouped_header_offsets_data_below_header(tmp_path):
    out = tmp_path / "g.xlsx"
    export_result_xlsx(_grouped_result(), str(out))
    ws = load_workbook(out)["Tablo"]
    # two header rows -> data starts at Excel row 3, and the pane freezes there
    assert ws["A3"].value == "a1"
    assert ws["C3"].value == "b1"
    assert ws.freeze_panes == "A3"


def test_export_grouped_header_highlights_data_disagreement_at_right_offset(tmp_path):
    result = _grouped_result()
    result["disagreements"] = [{"kind": "cell", "pos": (0, 4), "b1": "c2", "b2": "c9"}]
    out = tmp_path / "g.xlsx"
    export_result_xlsx(result, str(out))
    ws = load_workbook(out)["Tablo"]
    # data row 0, col 4 -> Excel E3 (below the 2-row header)
    assert ws["E3"].fill.patternType == "solid"
    assert ws["D3"].fill.patternType is None


def test_export_review_all_headers_highlights_whole_grouped_header(tmp_path):
    result = _grouped_result()
    result["review_all_headers"] = True          # undefined form -> flag whole header
    out = tmp_path / "g.xlsx"
    export_result_xlsx(result, str(out))
    ws = load_workbook(out)["Tablo"]
    assert ws["A1"].fill.patternType == "solid"       # top-level cell flagged
    assert ws["D2"].fill.patternType == "solid"       # sub cell flagged
    assert ws["A3"].fill.patternType is None          # data row not flagged


def _candidates_result():
    return {
        "mode": "consensus",
        "backends": ["b1", "b2"],
        "headers": [], "rows": [],
        "shape_match": False,
        "disagreements": [{"kind": "shape", "b1": (1, 3), "b2": (1, 4)}],
        "needs_review": True,
        "issues": ["modeller farkli sekil verdi"],
        "candidates": [
            {"backend": "b1", "headers": ["A", "B", "C"], "rows": [["1", "2", "3"]],
             "review_cells": {(0, 2)}, "suspect_count": 1, "number_fidelity": 0.67},
            {"backend": "b2", "headers": ["A", "B", "C", "D"], "rows": [["1", "2", "3", "4"]],
             "review_cells": set(), "suspect_count": 0, "number_fidelity": 1.0},
        ],
    }


def test_export_candidates_writes_comparison_and_per_model_sheets(tmp_path):
    out = tmp_path / "c.xlsx"
    export_result_xlsx(_candidates_result(), str(out))
    wb = load_workbook(out)
    assert wb.sheetnames == ["Karsilastirma", "Model_A_-_b1", "Model_B_-_b2", "Rapor"]
    assert [c.value for c in wb["Model_A_-_b1"][1]] == ["A", "B", "C"]
    assert [c.value for c in wb["Model_B_-_b2"][1]] == ["A", "B", "C", "D"]
    assert wb["Model_B_-_b2"]["D2"].value == "4"


def test_export_comparison_sheet_reports_per_model_summary(tmp_path):
    out = tmp_path / "c.xlsx"
    export_result_xlsx(_candidates_result(), str(out))
    ws = load_workbook(out)["Karsilastirma"]
    rows = {r[0].value: [c.value for c in r[1:]] for r in ws.iter_rows()}
    assert rows["Sutun sayisi"][:2] == [3, 4]
    assert rows["Supheli sayi (OCR'da yok)"][:2] == [1, 0]


def test_export_candidate_suspect_cell_highlighted_on_its_sheet(tmp_path):
    out = tmp_path / "c.xlsx"
    export_result_xlsx(_candidates_result(), str(out))
    wb = load_workbook(out)
    # b1's cell (0,2) flagged -> C2 amber; b2 has no suspects
    assert wb["Model_A_-_b1"]["C2"].fill.patternType == "solid"
    assert wb["Model_B_-_b2"]["D2"].fill.patternType is None


def test_export_single_backend_result_has_no_highlights(tmp_path):
    out = tmp_path / "t.xlsx"
    result = {
        "backend": "vl",
        "headers": ["Product", "Qty"],
        "rows": [["Pen", "5"]],
        "confidence": 1.0,
        "structural_confidence": 1.0,
        "number_fidelity": 1.0,
        "needs_review": False,
        "issues": [],
    }
    export_result_xlsx(result, str(out))
    ws = load_workbook(out)["Tablo"]
    assert ws["B2"].fill.patternType is None
