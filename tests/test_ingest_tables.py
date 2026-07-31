"""Ingest side of the table flow: how a pipeline result becomes a RAG chunk,
and that ingest extracts through the verified pipeline rather than a raw backend.
"""
import json

from pipeline.index import ingest as ingest_router
from pipeline.extraction import router


def _consensus_table():
    """Shaped like table_pipeline.run_consensus output."""
    return {
        "mode": "consensus",
        "backends": ["b1", "b2"],
        "headers": ["Kod", "Tutar"],
        "rows": [["1", "10,00"], ["2", "20,00"]],
        "confidence": 0.95,
        "agreement": 0.75,
        "structural_confidence": 1.0,
        "number_fidelity": 0.95,
        "shape_match": True,
        "disagreements": [{"kind": "cell", "pos": (1, 1), "b1": "20,00", "b2": "26,00"}],
        "issues": ["1 hucrede modeller ayristi (insan gozden gecirmeli)"],
        "needs_review": True,
    }


def _chunks(table, tmp_path, monkeypatch, source_tag="image:tables"):
    monkeypatch.setattr(ingest_router, "OUTPUT_DIR", tmp_path)
    return ingest_router.chunks_from_tables([table], source_tag, "doc")


# --- trust signals reach the chunk ---

def test_chunk_carries_consensus_metadata(tmp_path, monkeypatch):
    chunks = _chunks(_consensus_table(), tmp_path, monkeypatch)
    data = chunks[0]["table_data"]

    assert data["mode"] == "consensus"
    assert data["backends"] == ["b1", "b2"]
    assert data["agreement"] == 0.75
    assert data["disagreements"][0]["pos"] == (1, 1)
    assert data["number_fidelity"] == 0.95


def test_chunk_table_data_is_json_serializable(tmp_path, monkeypatch):
    """It is stored in a jsonb column, so a stray set/tuple would fail at insert."""
    chunks = _chunks(_consensus_table(), tmp_path, monkeypatch)
    assert json.dumps(chunks[0]["table_data"])


def test_pipeline_review_decision_is_respected_not_recomputed(tmp_path, monkeypatch):
    """The pipeline weighs agreement and number fidelity; ingest must not
    second-guess it with a confidence threshold it knows less about."""
    table = _consensus_table()
    table["confidence"] = 1.0          # would pass ingest's own threshold
    table["issues"] = []               # and carry no issues
    table["needs_review"] = True       # but the pipeline still says review

    chunks = _chunks(table, tmp_path, monkeypatch)
    assert chunks[0]["table_data"]["needs_review"] is True


def test_falls_back_to_shape_confidence_for_a_plain_table(tmp_path, monkeypatch):
    plain = {"headers": ["A", "B"], "rows": [["1", "2"]]}
    chunks = _chunks(plain, tmp_path, monkeypatch)
    data = chunks[0]["table_data"]

    assert data["confidence"] == 1.0
    assert data["needs_review"] is False
    assert "mode" not in data


# --- page numbering ---

def test_page_comes_from_the_source_tag():
    """The router converts ONE single-page PDF per page, so the parsed
    document always claims to be on page 1 -- the tag is the only place the real
    page number survives."""
    assert ingest_router.page_from_tag("page26:native") == 26
    assert ingest_router.page_from_tag("page7:scanned") == 7
    assert ingest_router.page_from_tag("page250:native") == 250


def test_page_is_zero_when_the_tag_carries_none():
    for tag in ("image:ocr", "image:tables", "", None):
        assert ingest_router.page_from_tag(tag) == 0


def test_table_chunk_takes_the_page_from_its_tag(tmp_path, monkeypatch):
    chunks = _chunks(_consensus_table(), tmp_path, monkeypatch, source_tag="page42:tables")
    assert chunks[0]["page"] == 42
    assert chunks[0]["table_data"]["page"] == 42


def test_plain_text_chunks_take_the_page_from_their_tag():
    chunks = ingest_router.chunk_plain_text("bir paragraf", "page9:scanned")
    assert chunks and all(c["page"] == 9 for c in chunks)


def test_table_text_has_no_embedded_citation_header(tmp_path, monkeypatch):
    """build_context adds a citation to every chunk from its metadata, so a
    header baked into the text would duplicate it and pollute the index."""
    text = _chunks(_consensus_table(), tmp_path, monkeypatch)[0]["text"]
    assert not text.startswith("Belge:")
    assert text.startswith("|")


# --- headings have to reach the index, not just the metadata ---

def test_headings_without_a_body_are_still_emitted():
    """A masthead is a run of headings with no text under it. Left at the
    default the chunker discards those entirely, so a document's own number and
    date -- among the most asked-about facts it has -- never reach the index."""
    assert ingest_router.chunker.always_emit_headings is True


def test_indexed_text_carries_the_heading_path(monkeypatch):
    """chunk.text drops headings, so anything that lives in one is unsearchable
    and a heading-only chunk indexes as an empty string."""
    class _Chunk:
        text = "govde"

    class _Chunker:
        def contextualize(self, chunk):
            return f"Baslik\n{chunk.text}"

    monkeypatch.setattr(ingest_router, "chunker", _Chunker())
    assert ingest_router.chunk_text(_Chunk()) == "Baslik\ngovde"


# --- fragments get folded back into their context ---

def _c(text, tag="page1:native", ctype="text", headings=None):
    return {"type": ctype, "text": text, "source_tag": tag, "page": 1,
            "headings": headings or []}


def test_a_fragment_joins_the_chunk_before_it():
    """The measured failure: a label/value fragment split away from the subject
    it describes, leaving nothing for a query to match on."""
    out = ingest_router.merge_small_chunks([_c("a" * 400), _c("b" * 40)], min_chars=150)
    assert len(out) == 1
    assert out[0]["text"] == "a" * 400 + "\n" + "b" * 40


def test_a_fragment_at_the_start_absorbs_the_next_chunk():
    """Merging only backwards would strand a fragment that comes first."""
    out = ingest_router.merge_small_chunks([_c("a" * 40), _c("b" * 400)], min_chars=150)
    assert len(out) == 1


def test_consecutive_fragments_fold_together():
    out = ingest_router.merge_small_chunks(
        [_c("a" * 40), _c("b" * 40), _c("c" * 40)], min_chars=150)
    assert len(out) == 1


def test_page_boundary_is_never_crossed():
    """A merge across pages would file text under the wrong page number, which
    is exactly the citation error this pipeline just finished fixing."""
    out = ingest_router.merge_small_chunks(
        [_c("a" * 400, tag="page1:native"), _c("b" * 40, tag="page2:native")],
        min_chars=150)
    assert len(out) == 2


def test_a_table_fragment_is_not_folded_into_prose():
    out = ingest_router.merge_small_chunks(
        [_c("a" * 400, ctype="text"), _c("b" * 40, ctype="table")], min_chars=150)
    assert len(out) == 2


def test_merging_stops_at_the_length_ceiling():
    out = ingest_router.merge_small_chunks(
        [_c("a" * 400), _c("b" * 40)], min_chars=150, max_chars=300)
    assert len(out) == 2


def test_normal_sized_chunks_are_left_alone():
    out = ingest_router.merge_small_chunks([_c("a" * 400), _c("b" * 400)], min_chars=150)
    assert len(out) == 2


def test_the_survivor_inherits_headings_when_it_has_none():
    out = ingest_router.merge_small_chunks(
        [_c("a" * 400), _c("b" * 40, headings=["Bolum"])], min_chars=150)
    assert out[0]["headings"] == ["Bolum"]


def test_input_chunks_are_not_mutated():
    """main() reuses the list it passes in; silently editing it would make the
    merge count wrong and the behaviour order-dependent."""
    original = [_c("a" * 400), _c("b" * 40)]
    ingest_router.merge_small_chunks(original, min_chars=150)
    assert original[0]["text"] == "a" * 400


# --- exports ---

def test_flagged_table_is_copied_to_review_with_a_report(tmp_path, monkeypatch):
    _chunks(_consensus_table(), tmp_path, monkeypatch)
    report = tmp_path / "tables" / "_review" / "doc_image_tables_0.issues.txt"

    assert (tmp_path / "tables" / "_review" / "doc_image_tables_0.xlsx").exists()
    text = report.read_text(encoding="utf-8")
    assert "b1+b2" in text
    assert "1 hucrede ayrisma" in text


def test_pipeline_result_gets_the_rich_export(tmp_path, monkeypatch):
    """A consensus result has disagreements to highlight, so it must go through
    export_result_xlsx (Tablo + Rapor sheets), not the plain writer."""
    from openpyxl import load_workbook

    _chunks(_consensus_table(), tmp_path, monkeypatch)
    wb = load_workbook(tmp_path / "tables" / "doc_image_tables_0.xlsx")
    assert "Rapor" in wb.sheetnames


def test_plain_table_gets_the_plain_export(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    _chunks({"headers": ["A"], "rows": [["1"]]}, tmp_path, monkeypatch)
    wb = load_workbook(tmp_path / "tables" / "doc_image_tables_0.xlsx")
    assert "Rapor" not in wb.sheetnames


# --- ingest extracts through the pipeline, not a raw backend ---

def test_ingest_uses_consensus_by_default(monkeypatch):
    seen = {}

    def fake_run_consensus(image_path, ocr_text=None):
        seen["mode"] = "consensus"
        seen["ocr_text"] = ocr_text
        return [{"headers": ["A"], "rows": [["1"]]}]

    import pipeline.extraction.table_pipeline as tp
    monkeypatch.setattr(tp, "run_consensus", fake_run_consensus)
    monkeypatch.setattr(router, "INGEST_TABLE_MODE", "consensus")

    out = router.tables_from_image_verified("x.png", ocr_text="onceden okundu")
    assert seen["mode"] == "consensus"
    # the caller's OCR reading is reused instead of re-running the service
    assert seen["ocr_text"] == "onceden okundu"
    assert out[0]["headers"] == ["A"]


def test_ingest_single_mode_runs_one_backend(monkeypatch):
    seen = {}

    def fake_run(image_path, ocr_text=None):
        seen["mode"] = "single"
        return []

    import pipeline.extraction.table_pipeline as tp
    monkeypatch.setattr(tp, "run", fake_run)
    monkeypatch.setattr(router, "INGEST_TABLE_MODE", "single")

    router.tables_from_image_verified("x.png", ocr_text="t")
    assert seen["mode"] == "single"
