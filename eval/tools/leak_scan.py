"""The leak-scan battery as a tracked tool, with its scars written into code.

Every rule below was learned by a scan that FAILED first:

  * The corpus is built by ENUMERATING data/ -- never from a list. A stale
    list once held 2 of 8 documents and cleared two leaks.
  * TWO haystacks, because "absent from the document text layer" is NOT
    clearance: a real leak lived in table content that reaches the corpus
    through EXTRACTION, so the FULL haystack also folds in our saved
    outputs under output/, where extraction results live.
  * A reader that cannot read is a scan that cannot clear. If PDFs exist
    and none contributed text, or spreadsheets exist and none could be
    opened, the scan RAISES -- the first version swallowed a missing PDF
    library and blessed eight unread documents.
  * CHARACTER windows, direction inverted: repo windows are collected once
    into a set and each corpus is walked ONCE against it. Word windows
    broke on punctuation stuck to a token, and a fixed 14-character floor
    once excluded the 13-character leak being hunted -- width stays a
    parameter.
  * Post-commit scans take their file list FROM THE COMMIT (`git show`),
    never from `git diff`, which is empty after committing and scans
    nothing while reporting clean.
  * A scan that CRASHES or scans ZERO files is a FAILED scan, never a
    clean one.
  * A secret is reported as a LABEL and a location, never as its text: an
    earlier version echoed the first 40 characters of what it caught,
    putting password material into the very report meant to keep it out.
    Known placeholders are skipped so the tool does not cry wolf forever
    -- but a DSN's clearance is judged on its PASSWORD FIELD alone (a real
    password containing "..." was once cleared by the dots), and tests/
    fixtures are TRIAGED rather than exempted (a real credential pasted
    into a test was once invisible by path).
  * The tool cannot certify what it cannot read: source images, an
    unscannable output slice, or a tracked repo file with an unsupported
    extension each cap the verdict at TRIAGE. Data-side spreadsheets are
    read whole -- the tool-sheet skip applies only to our own output tree,
    because a source sheet wearing a skippable name is still document
    content.
  * A source-name piece that is also repo-filename vocabulary is reported
    for triage, never dropped: the old exclusion let a repo file named
    after the source document allowlist its own name out of the battery.

The tool itself must carry no document-derived text; its own test feeds it
a synthetic corpus and proves the scan finds a planted fragment. It is one
layer of the battery, not the battery: an independent scan remains part of
every release verdict.
"""
import argparse
import re
import sqlite3
import subprocess
import sys
import unicodedata
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]

TEXT_SUFFIXES = {".txt", ".md", ".json", ".jsonl", ".csv", ".py", ".yml",
                 ".yaml", ".toml", ".cfg", ".sql", ".sh", ".ini", ".pem",
                 ".example", ".env", ""}
# Image files cannot be text-scanned; they are COUNTED and declared, never
# silently skipped -- the image gap is a known limit closed by human eyes,
# and a limit that stops being printed stops being known.
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff"}
INVISIBLE = {"\u00a0": "NBSP", "\u202f": "NNBSP", "\u200b": "ZWSP",
             "\ufeff": "BOM",
             # direction-override characters can make displayed source lie
             # about what executes; none has a legitimate place in this repo
             "\u202a": "BIDI", "\u202b": "BIDI", "\u202c": "BIDI",
             "\u202d": "BIDI", "\u202e": "BIDI", "\u2066": "BIDI",
             "\u2067": "BIDI", "\u2068": "BIDI", "\u2069": "BIDI"}
SECRETS = {
    "dsn_kimlikli": re.compile(r"(?i)postgres(?:ql)?://\S+:\S+@"),
    "aws_anahtari": re.compile(r"AKIA[0-9A-Z]{16}"),
    # the private-key HEADER, not the bare marker: the bare "-----BEGIN"
    # matched this very dictionary's own source, and a certificate is not
    # a credential -- keys (encrypted ones included) are
    "pem_anahtari": re.compile(
        r"-----BEGIN (?:ENCRYPTED )?(?:RSA |EC |DSA |OPENSSH )?"
        r"PRIVATE KEY-----"),
    "sk_anahtari": re.compile(r"sk-[A-Za-z0-9]{20,}"),
    "bearer_jetonu": re.compile(r"Bearer\s+[A-Za-z0-9._-]{16,}"),
    "api_anahtari": re.compile(r"(?i)api[_-]?key\s*[:=]\s*['\"][^'\"]{8,}"),
}
# DECLARED STATIC-SCAN LIMIT: a credential assembled at runtime from pieces
# ("postgres" + "://...") is invisible to every pattern above. That is a
# property of static scanning, not a bug this file can fix wholesale; the
# high-entropy triage layer below narrows it, and the review convention
# that secrets never enter source in any form remains the real mitigation
# (an independent scanner such as gitleaks stays part of the battery).
#
# A placeholder clears a match ONLY when the placeholder token covers the
# WHOLE credential core for that label. Two probes drove this shape: a
# real password CONTAINING "..." was cleared by the dots, and a
# real-looking token CARRYING a placeholder word anywhere in the match
# was cleared by substring search. Labels with no extractor (aws/sk/pem
# keys) have no core a placeholder could stand in for: nothing clears
# them.
_SECRET_CORES = {
    "dsn_kimlikli": re.compile(r"://[^:/@\s]*:([^@\s]+)@"),
    "api_anahtari": re.compile(r"['\"]([^'\"]{8,})"),
    "bearer_jetonu": re.compile(r"(?i)Bearer\s+(\S+)"),
}
_PLACEHOLDER_TOKEN = re.compile(
    r"^(?:CHANGE_ME|DEGISTIR|\.{3,}|<[^>]{1,40}>|Enter_Your[\w-]*|"
    r"[\w-]*_Here|yer_tutucu[\w-]*)$",
    re.IGNORECASE)


def _documented_placeholder(label: str, matched: str) -> bool:
    extractor = _SECRET_CORES.get(label)
    if extractor is None:
        return False
    core = extractor.search(matched)
    token = core.group(1) if core else matched
    return bool(_PLACEHOLDER_TOKEN.match(token))


def _masked_name(path_like) -> str:
    """A path reduced to a masked basename: error output must locate a
    problem file without republishing a data-derived name -- a probe
    showed broken source/output paths echoing whole into the report."""
    p = Path(str(path_like))
    stem = p.stem
    masked = (stem if len(stem) <= 2
              else stem[0] + "*" * (len(stem) - 2) + stem[-1])
    return masked + p.suffix.lower()


# High-entropy string literals: the triage answer to PIECEWISE secrets. A
# quoted token this long and this uniform is a key-shaped object whatever
# its variable name says; 4.2 bits/char sits above pure-hex uniformity
# (4.0 exactly), so hash pins do not fire while base64-class material
# does. Reported as counts and lines, never text.
_ENTROPY_TOKEN = re.compile(r"['\"]([A-Za-z0-9+/=_\-]{20,})['\"]")
_ENTROPY_FLOOR = 4.2


def _shannon(text: str) -> float:
    import math

    counts: dict[str, int] = {}
    for ch in text:
        counts[ch] = counts.get(ch, 0) + 1
    total = len(text)
    return -sum((c / total) * math.log2(c / total)
                for c in counts.values())


def fold(text: str) -> str:
    """Case/diacritic folding, self-contained on purpose: importing the
    pipeline's fold would make the scanner unusable exactly when the
    pipeline is broken, which is when it is needed most.

    The dotted/dotless i pairs are mapped BEFORE lowercasing: "I".lower()
    is ASCII "i" but Turkish "I" is dotless-i's capital, and relying on
    lower() alone left one direction of the pair unfolded."""
    s = str(text)
    for source, target in (("İ", "i"), ("I", "i"), ("ı", "i"), ("ý", "i")):
        s = s.replace(source, target)
    s = s.lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


def _read_pdf(path: Path) -> str | None:
    """PyMuPDF first (installed here), pypdf as fallback; None = unreadable.

    A VALID pdf with no text layer also returns None: "opened fine, gave
    nothing" is exactly how a scanned document hides from a text scan, and
    counting it as a successful contribution once blessed a corpus that
    held none of its content."""
    body = None
    try:
        import fitz

        with fitz.open(str(path)) as doc:
            body = " ".join(page.get_text() for page in doc)
    except Exception:
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            body = " ".join(
                page.extract_text() or "" for page in reader.pages)
        except Exception:
            return None
    if body is None or len(body.strip()) < 20:
        return None
    return body


def _read_xlsx(path: Path, skip_tool_sheets: bool = True) -> str | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(str(path), read_only=True, data_only=True)
        cells = []
        for sheet in workbook.worksheets:
            # report/comparison sheets in OUR OUTPUT tree carry the tool's
            # own text (review messages, model comparisons), not document
            # content; indexing them made the scanner flag the very tests
            # that assert those messages. The skip applies ONLY to the
            # output side: a SOURCE workbook's sheet wearing one of these
            # names is document content like any other, and skipping it by
            # title once excluded real content from the haystack.
            folded_title = fold(sheet.title)
            if skip_tool_sheets and ("rapor" in folded_title
                                     or "karsilastirma" in folded_title):
                continue
            for row in sheet.iter_rows(values_only=True):
                cells.extend(str(cell) for cell in row if cell is not None)
        workbook.close()
        return " ".join(cells)
    except Exception:
        return None


def _read_sqlite(path: Path) -> str | None:
    try:
        connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        dump = " ".join(connection.iterdump())
        connection.close()
        return dump
    except Exception:
        return None


def _read_json_values(path: Path) -> str | None:
    """JSON leaf VALUES only, keys skipped. The corpus-side JSON is our own
    format: ground truth, question sets, saved answers. Their VALUES are
    document-derived content and belong in the haystack; their KEYS are our
    schema vocabulary, and indexing them made every repo file that spells
    its own field names a permanent false hit."""
    import json as _json

    try:
        body = path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return None
    try:
        data = _json.loads(body)
    except ValueError:
        return body  # not valid JSON: raw text is better than nothing
    values = []

    def walk(node):
        if isinstance(node, dict):
            for item in node.values():
                walk(item)
        elif isinstance(node, list):
            for item in node:
                walk(item)
        elif node is not None:
            values.append(str(node))

    walk(data)
    return " ".join(values)


_READERS = {
    ".pdf": _read_pdf,
    ".xlsx": _read_xlsx,
    ".xlsm": _read_xlsx,
    ".db": _read_sqlite,
    ".sqlite": _read_sqlite,
    ".json": _read_json_values,
}


def _collect(root: Path, tool_sheets: bool = True):
    """(folded_text, contributed, unreadable, images, unknown) per file.

    Every enumerated file lands in exactly one bucket. "I did not know what
    that was" and "that was an image" are RETURNED, because a category the
    caller cannot see is a category that silently shrinks the corpus -- 12
    of 34 data files were once outside the haystack with nothing printed.
    ``tool_sheets=False`` (the data/ side) reads EVERY worksheet."""
    pieces, contributed, unreadable, images, unknown = [], 0, [], [], []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        reader = _READERS.get(suffix)
        if reader is not None:
            if reader is _read_xlsx:
                body = reader(path, skip_tool_sheets=tool_sheets)
            else:
                body = reader(path)
            if body is None:
                unreadable.append(str(path))
            else:
                pieces.append(body)
                contributed += 1
        elif suffix in TEXT_SUFFIXES:
            try:
                pieces.append(path.read_text(encoding="utf-8",
                                             errors="ignore"))
                contributed += 1
            except OSError:
                unreadable.append(str(path))
        elif suffix in IMAGE_SUFFIXES:
            images.append(str(path))
        else:
            unknown.append(str(path))
    return fold("\n".join(pieces)), contributed, unreadable, images, unknown


def build_corpora(data_dir: Path, output_dir: Path | None):
    """The document haystack (data/) and the full haystack (+ output/).

    An unreadable file under data/ is a HARD ERROR: every skipped document
    is a place a leak can hide, and the first version skipped all eight
    PDFs in silence. Under output/ the rule is VISIBILITY plus triage,
    not death: our own artifact tree accumulates audit litter (pytest temp
    dirs carrying fake fixture "PDFs"), and a scan that dies on its own
    byproducts is a scan that gets skipped -- the historic failure. But
    what could not join the haystack is COUNTED AND RETURNED (unreadable,
    images, unknown types alike): a probe showed output-side images and
    unknown files being dropped with nothing printed, and the full-only
    verdicts wearing more certainty than the corpus carried."""
    documents, contributed, unreadable, images, unknown = _collect(
        data_dir, tool_sheets=False)
    if unreadable:
        raise RuntimeError(
            f"korpus eksik, tarama BASARISIZ: okunamayan {len(unreadable)} "
            f"dosya: {[_masked_name(p) for p in unreadable[:5]]}")
    if unknown:
        raise RuntimeError(
            f"korpus eksik, tarama BASARISIZ: taninmayan tur "
            f"{len(unknown)} dosya: {[_masked_name(p) for p in unknown[:5]]}")
    if contributed == 0:
        raise RuntimeError("korpus bos: data/ hic katki vermedi -- tarama "
                           "temiz raporlayamaz")
    full = documents
    output_unreadable, output_images, output_unknown = [], [], []
    if output_dir is not None and output_dir.is_dir():
        saved, _, output_unreadable, output_images, output_unknown = (
            _collect(output_dir))
        full = documents + "\n" + saved
    # SOURCE-DOCUMENT name pieces travel with the corpora: code wearing a
    # source file's name piece is a reference leak even when no content
    # matches -- the audited case was two source PDFs' year stems in test
    # fixtures. Narrowings learned on the live tree: only pdf/image stems
    # count (our own eval artifacts' names -- ordinary words like the
    # question-set names -- flagged the entire repo), with digit pieces
    # identifying at 4 characters and words at 7. A piece that is ALSO
    # part of a tracked repo file's name is NOT silently dropped any more:
    # that exclusion let a repo file named after the source document
    # allowlist its own name into invisibility (and made the filename
    # layer unreachable by construction). Overlaps are returned as their
    # own set and reported for human triage instead.
    import re as _re

    repo_name_pieces = set()
    for name in tracked_files():
        for piece in _re.split(r"[_\-. /\\]+", name.lower()):
            if piece:
                repo_name_pieces.add(fold(piece))
    hard_stems, vocab_stems = set(), set()

    def classify(token):
        floor = 4 if token.isdigit() else 7
        if len(token) < floor:
            return
        if token in repo_name_pieces:
            vocab_stems.add(token)
        else:
            hard_stems.add(token)

    source_suffixes = {".pdf"} | IMAGE_SUFFIXES
    for path in data_dir.rglob("*"):
        if path.is_file() and path.suffix.lower() in source_suffixes:
            pieces = [fold(piece)
                      for piece in _re.split(r"[_\-. ]+", path.stem)
                      if piece]
            for piece in pieces:
                classify(piece)
            # the whole multi-piece name, re-joined every common way: a
            # probe wrote the name with a DIFFERENT separator than the
            # file uses ("a_b" for "a-b") and no single piece cleared the
            # floor -- the joined variants carry the identity the pieces
            # alone lose
            if len(pieces) > 1:
                for joiner in ("", "_", "-", " ", "."):
                    classify(joiner.join(pieces))
    return (documents, full, contributed, output_unreadable,
            sorted(hard_stems), sorted(vocab_stems), len(images),
            output_images, output_unknown)


def _git_lines(args) -> list[str]:
    """NUL-separated git listings: `.split()` on plain output silently
    dropped every path containing a space -- files the scan then never
    opened while reporting clean."""
    out = subprocess.run(args + ["-z"], capture_output=True, text=True,
                         cwd=REPO_ROOT, check=True).stdout
    return [name for name in out.split("\0") if name]


def tracked_files(commits=None) -> list[str]:
    """The scan population: tracked+changed by default, or the files OF THE
    GIVEN COMMITS -- the post-commit mode that `git diff` cannot provide."""
    if commits:
        names = set()
        for commit in commits:
            names.update(_git_lines(
                ["git", "show", "--name-only", "--format=", commit]))
        return sorted(names)
    tracked = _git_lines(["git", "ls-files"])
    untracked = _git_lines(["git", "ls-files", "-o", "--exclude-standard"])
    return sorted(set(tracked) | set(untracked))


def commit_messages(commits) -> list[tuple[str, str]]:
    """Commit MESSAGES are scan input too: a message can carry a document
    phrase or a secret as easily as a file, and the historic battery
    scanned them while the first version of this tool did not."""
    out = []
    for commit in commits or ():
        message = subprocess.run(
            ["git", "log", "-1", "--format=%B", commit],
            capture_output=True, text=True, cwd=REPO_ROOT, check=True,
        ).stdout
        out.append((f"commit-mesaji:{commit[:12]}", message))
    return out


def read_repo_file(name: str, commit: str | None = None) -> str | None:
    """File content from the git object when a commit is named, else the
    working tree -- scanning what was COMMITTED, not what remains."""
    if commit:
        result = subprocess.run(["git", "show", f"{commit}:{name}"],
                                capture_output=True, text=True,
                                encoding="utf-8", errors="ignore",
                                cwd=REPO_ROOT)
        return result.stdout if result.returncode == 0 else None
    path = REPO_ROOT / name
    if not path.is_file() or path.suffix.lower() not in TEXT_SUFFIXES:
        return None
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return None


def _locate_lines(body: str, window: str) -> list:
    """Line numbers whose FOLDED text contains the window -- the report
    points AT the leak instead of repeating it. A window that only exists
    across a collapsed line break locates as line 0."""
    found = []
    for number, line in enumerate(body.splitlines(), 1):
        if window in fold(line):
            found.append(number)
            if len(found) == 3:
                return found
    return found or [0]


def scan(files, documents: str, full: str, width: int,
         commit: str | None = None, extra_bodies=None,
         data_stems=None, vocab_stems=None) -> dict:
    """One pass over each haystack against a window set built ONCE.

    Repo text is the short side: all its windows go into one dict
    (window -> files), then each corpus is walked a single time. Document
    hits are leaks; full-only hits are our own saved output echoing back
    and get triaged, not ignored.

    HITS ARE REPORTED AS LOCATIONS -- file, window count, line numbers --
    never as the matched text. An earlier version printed the windows
    themselves, which put document fragments into the very report built to
    prove their absence.

    A repo file the scan CANNOT read (unsupported suffix, in the working
    tree) lands in ``kapsam_disi``: it used to be silently excluded, which
    let any binary-shaped file opt out of the battery by its extension."""
    if not files:
        raise RuntimeError("tarama girdisi bos: bu TEMIZ degil, BASARISIZDIR")
    report = {"files": 0, "document_hits": {}, "document_prose_hits": {},
              "full_only_hits": {}, "secrets": {}, "test_secrets": {},
              "invisible": {}, "stem_hits": {}, "stem_vocab_hits": {},
              "yuksek_entropi": {}, "kapsam_disi": []}
    windows: dict[str, set] = {}
    bodies: dict[str, str] = {}
    stems = [s for s in (data_stems or ()) if len(s) >= 4]
    vocab = [s for s in (vocab_stems or ()) if len(s) >= 4]

    def _masked(piece):
        return piece[0] + "*" * (len(piece) - 2) + piece[-1]

    def _stem_search(piece, folded):
        return re.search(r"(?<![a-z0-9])" + re.escape(piece)
                         + r"(?![a-z0-9])", folded)

    def take(name, body):
        report["files"] += 1
        bodies[name] = body
        folded = fold(body)
        for i in range(max(len(folded) - width + 1, 0)):
            windows.setdefault(folded[i:i + width], set()).add(name)
        # a fixture-marked path routes its secret-pattern hits to a TRIAGE
        # bucket by PATH, not by vocabulary. Round 17: the earlier version
        # exempted tests/ entirely, and a real credential pasted into a
        # test would have been invisible; now it is listed and a human
        # says "fixture" out loud.
        fixture_path = name.startswith("tests/") or name.startswith("tests\\")
        labels = [
            label for label, pattern in SECRETS.items()
            if any(not _documented_placeholder(label, m.group())
                   for m in pattern.finditer(body))
        ]
        if labels:
            bucket = "test_secrets" if fixture_path else "secrets"
            report[bucket][name] = sorted(labels)  # labels, NEVER text
        ghosts = sorted({label for ch, label in INVISIBLE.items()
                         if ch in body})
        if ghosts:
            report["invisible"][name] = ghosts
        # data file-name pieces INSIDE content: the audited case was source
        # files' year stems in fixtures. Reported masked -- the stem itself
        # is data-derived and must not ride out in the report. Pieces that
        # are ALSO repo-filename vocabulary go to their own triage bucket:
        # dropping them entirely once let a repo file named after the
        # source document allowlist the name out of the battery.
        found = sorted(_masked(piece) for piece in stems
                       if _stem_search(piece, folded))
        if found:
            report["stem_hits"][name] = found[:8]
        overlap = sorted(_masked(piece) for piece in vocab
                         if _stem_search(piece, folded))
        if overlap:
            report["stem_vocab_hits"][name] = overlap[:8]
        # quoted key-shaped tokens: counts and lines, never the token --
        # the triage net under piecewise/assembled secrets
        entropic_lines = set()
        hits = 0
        for match in _ENTROPY_TOKEN.finditer(body):
            if _shannon(match.group(1)) > _ENTROPY_FLOOR:
                hits += 1
                line = body.count("\n", 0, match.start()) + 1
                entropic_lines.add(line)
        if hits:
            report["yuksek_entropi"][name] = {
                "adet": hits, "satirlar": sorted(entropic_lines)[:6]}

    skipped = []
    for name in files:
        # ONE suffix policy for both modes: commit mode used to git-show
        # anything as text, and a commit holding only binary files scanned
        # its own garbage "cleanly" to exit 0. What the battery cannot
        # read, in either mode, is recorded out-of-scope -- never skipped
        # in silence, never counted as scanned.
        suffix = Path(name).suffix.lower()
        if suffix not in TEXT_SUFFIXES:
            if commit is not None or (REPO_ROOT / name).is_file():
                skipped.append(name)
            continue
        body = read_repo_file(name, commit)
        if body is not None:
            take(name, body)
        elif (commit is not None) or (REPO_ROOT / name).is_file():
            skipped.append(name)
    for name, body in (extra_bodies or ()):
        take(name, body)
    report["kapsam_disi"] = sorted(skipped)
    if report["files"] == 0:
        raise RuntimeError(
            "hicbir dosya okunamadi: bu TEMIZ degil, BASARISIZDIR")

    def sweep(corpus):
        found = {}
        for i in range(max(len(corpus) - width + 1, 0)):
            piece = corpus[i:i + width]
            if piece in windows:
                found.setdefault(piece, sorted(windows[piece]))
        return found

    in_documents = sweep(documents)
    in_full = sweep(full)
    # What carries a document FACT is a MULTI-DIGIT RUN that is not all
    # zeros. Three earlier splits over-fired in turn: punctuation flagged
    # JSON-building code (shared grammar), single digits flagged counting
    # grammar ("2 satir"), and bare "00" flagged every zero-filled grid a
    # test fixture shares with a transcribed empty table -- a zero is
    # formatting, not a fact. Everything below the bar goes to triage: an
    # obligation, never an acquittal.
    def _is_specific(piece):
        return any(set(run) != {"0"}
                   for run in re.findall(r"\d{2,}", piece))
    located: dict[str, dict] = {"document_hits": {},
                                "document_prose_hits": {},
                                "full_only_hits": {}}
    for piece, names in in_documents.items():
        bucket = ("document_hits" if _is_specific(piece)
                  else "document_prose_hits")
        for name in names:
            entry = located[bucket].setdefault(
                name, {"pencere": 0, "satirlar": set()})
            entry["pencere"] += 1
            entry["satirlar"].update(_locate_lines(bodies[name], piece))
    for piece, names in in_full.items():
        if piece in in_documents:
            continue
        for name in names:
            entry = located["full_only_hits"].setdefault(
                name, {"pencere": 0, "satirlar": set()})
            entry["pencere"] += 1
            entry["satirlar"].update(_locate_lines(bodies[name], piece))
    for bucket, entries in located.items():
        report[bucket] = {
            name: {"pencere": entry["pencere"],
                   "satirlar": sorted(entry["satirlar"])[:6]}
            for name, entry in sorted(entries.items())
        }
    return report


def main():
    parser = argparse.ArgumentParser(
        description="Sizinti taramasi: korpus data/ NUMARALANDIRILARAK "
                    "kurulur; okunamayan dosya/cokme/sifir-dosya = "
                    "basarisizlik")
    parser.add_argument("--commits", nargs="*", default=None,
                        help="verilirse dosya listesi ve icerik bu "
                             "commit'lerden alinir, diff'ten degil")
    parser.add_argument("--width", type=int, default=14,
                        help="karakter penceresi; supheli bir sizintiyi "
                             "avlarken ONCE onun boyuna indir")
    parser.add_argument("--data-dir", default=None)
    parser.add_argument("--output-dir", default=None,
                        help="tam samanligin ikinci yarisi (varsayilan "
                             "output/); bos gecmek yalniz belge samanligi "
                             "birakir")
    args = parser.parse_args()

    if not (8 <= args.width <= 64):
        raise SystemExit("pencere 8-64 araliginda olmali: cok dar her seyi, "
                         "cok genis hicbir seyi eslestirir -- ikisi de "
                         "false-clean uretir")
    data_dir = Path(args.data_dir) if args.data_dir else REPO_ROOT / "data"
    if not data_dir.is_dir():
        raise SystemExit(f"veri dizini yok: {data_dir} -- korpussuz tarama "
                         f"temiz raporlayamaz")
    output_dir = (Path(args.output_dir) if args.output_dir
                  else REPO_ROOT / "output")
    (documents, full, contributed, output_unreadable, data_stems,
     vocab_stems, image_count, output_images, output_unknown) = (
        build_corpora(data_dir, output_dir))
    files = tracked_files(args.commits)
    print(f"belge samanligi {len(documents):,} krk ({contributed} dosya) · "
          f"tam samanlik {len(full):,} krk · taranacak {len(files)} dosya · "
          f"pencere {args.width}")
    if image_count:
        # an unscannable slice of the SOURCE corpus: the tool must not say
        # TEMIZ about content it never read -- triage, closed by human eyes
        print(f"BILINEN SINIR: {image_count} kaynak gorsel metin-taranamaz; "
              f"insan gozuyle kapatilana kadar tarama en iyi ihtimalle "
              f"TRIYAJDIR")
    if output_unreadable:
        print(f"UYARI: cikti samanligina katilamayan {len(output_unreadable)} "
              f"dosya (ilk 3, maskeli): "
              f"{[_masked_name(p) for p in output_unreadable[:3]]}")
    if output_images or output_unknown:
        print(f"UYARI: cikti tarafinda taranamayan {len(output_images)} "
              f"gorsel, {len(output_unknown)} taninmayan dosya -- yalniz-tam"
              f"-samanlik hukumleri bu bosluk kadar eksiktir")

    # staged-but-uncommitted content differs from the worktree exactly when
    # it matters; it is scanned as its own virtual file
    extra = commit_messages(args.commits) if args.commits else []
    if not args.commits:
        for name in _git_lines(["git", "diff", "--cached", "--name-only"]):
            staged = subprocess.run(
                ["git", "show", f":{name}"], capture_output=True, text=True,
                encoding="utf-8", errors="ignore", cwd=REPO_ROOT)
            if staged.returncode == 0:
                extra.append((f"{name} (staged)", staged.stdout))

    buckets = ("document_hits", "document_prose_hits", "full_only_hits",
               "secrets", "test_secrets", "invisible", "stem_hits",
               "stem_vocab_hits", "yuksek_entropi")
    if args.commits and len(args.commits) > 1:
        merged = {key: {} for key in buckets}
        merged["files"] = 0
        merged["kapsam_disi"] = []
        for one in args.commits:
            partial = scan(tracked_files([one]), documents, full,
                           args.width, one,
                           extra_bodies=commit_messages([one]),
                           data_stems=data_stems, vocab_stems=vocab_stems)
            merged["files"] += partial["files"]
            for key in buckets:
                merged[key].update(partial[key])
        report = merged
    else:
        commit = args.commits[0] if args.commits else None
        report = scan(files, documents, full, args.width, commit,
                      extra_bodies=extra, data_stems=data_stems,
                      vocab_stems=vocab_stems)

    # a repo file wearing a data file's name: reference leak, triage
    # bucket. ALL stems count here -- the repo-vocabulary overlap set most
    # of all, because "the repo already has a file by that name" is the
    # very situation this layer exists to catch.
    all_stems = set(data_stems) | set(vocab_stems)
    name_hits = sorted({
        name for name in files
        if len(Path(name).stem) >= 5 and fold(Path(name).stem) in all_stems
    })

    print(f"taranan dosya: {report['files']}")
    hard = triage = False
    for key, label, is_hard in (
            ("document_hits", "BELGE isabeti (rakam/sembol tasiyan)", True),
            ("secrets", "sir deseni", True),
            ("invisible", "gorunmez karakter", True),
            ("stem_hits", "veri dosya-adi parcasi ICERIKTE (maskeli)", True),
            ("test_secrets",
             "sir deseni tests/ altinda (fikstur mu, insan soylesin)", False),
            ("stem_vocab_hits",
             "veri dosya-adi parcasi ICERIKTE (repo sozlugunde de var; "
             "maskeli)", False),
            ("yuksek_entropi",
             "yuksek entropili dizgi (parcali/birlesik sir agi; sayi+satir)",
             False),
            ("document_prose_hits",
             "belge duz-metin eslesmesi (siradan soz dagarcigi)", False),
            ("full_only_hits", "yalniz-tam-samanlik isabeti", False)):
        if report[key]:
            hard = hard or is_hard
            triage = triage or not is_hard
            print(f"!! {label}:")
            for name, hits in sorted(report[key].items()):
                print(f"   {name}: {hits}")
        else:
            print(f"{label}: 0")
    if name_hits:
        triage = True
        print(f"!! veri dosya adi tasiyan depo dosyasi: {name_hits}")
    else:
        print("veri dosya adi eslesmesi: 0")
    if report.get("kapsam_disi"):
        # a tracked file the battery cannot read is an unscanned surface,
        # not a free pass
        triage = True
        print(f"!! taranamayan depo dosyasi ({len(report['kapsam_disi'])}): "
              f"{report['kapsam_disi'][:6]}")
    else:
        print("taranamayan depo dosyasi: 0")
    if output_unreadable or output_images or output_unknown:
        # an unreadable slice of the FULL haystack means the full-only
        # verdicts are incomplete: never clean, always at least triage
        triage = True
    if image_count:
        # unscannable SOURCE content: no exit-0 while it exists
        triage = True

    # Three-way exit, because "TEMIZ" next to a list of candidates trained
    # people to stop reading: hard findings fail, outstanding triage says
    # so in both the text and the exit code, and only silence is clean.
    if hard:
        print("SERT BULGU: tarama BASARISIZ")
        sys.exit(1)
    if triage:
        print("TRIYAJ GEREKLI: yukaridaki adaylar insan hukmu bekliyor; "
              "bu cikti TEMIZ degildir")
        sys.exit(2)
    print("TEMIZ")


if __name__ == "__main__":
    main()
